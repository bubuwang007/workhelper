from typing import Sequence
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from workhelper.excel.styles import text_length

class WorkSheet(Worksheet):

    def cell(
        self,
        row: int,
        column: int,
        value = None,
        format = None,
        alignment = None,
        font = None,
    ) -> Cell:
        cl = super().cell(row, column, value)
        if alignment is not None:
            cl.alignment = alignment
        if format is not None:
            cl.number_format = format
        if font is not None:
            cl.font = font
        return cl

    def auto_width(self, min_width: int | float = 9) -> None:
        for i in self.iter_cols():
            max_length = 0
            column = i[0].column
            for cell in i:
                if cell.value is not None:
                    length = text_length(cell)
                    max_length = max(max_length, length)
            max_length = max(max_length, min_width)
            self.column_dimensions[get_column_letter(column)].width = max_length

    def get_list(self) -> list[list]:
        ret = []
        for row in self.iter_rows():
            if row[0].value is None:
                continue
            tmp = []
            for cell in row:
                tmp.append(cell.value)
            while tmp[-1] is None:
                tmp.pop()
            ret.append(tmp)
        return ret

    def get_dict(self) -> dict:
        ret = {}
        li = self.get_list()
        for i in li:
            if len(i) == 0:
                ret.update({i[0]: None})
            elif len(i) == 2:
                ret.update({i[0]: i[1]})
            else:
                ret.update({i[0]: i[1:]})
        return ret

    def set_list(
        self,
        li: list,
        start_row: int = 1,
        start_column: int = 1,
        format = None,
        alignment = None,
        font = None,
        direction: str = 'row',
    ) -> None:
        if direction == 'row':
            for i, cell in enumerate(li):
                    self.cell(
                        start_row + i,
                        start_column,
                        cell,
                        format,
                        alignment,
                        font
                    )
        elif direction == 'col':
            for i, cell in enumerate(li):
                    self.cell(
                        start_row,
                        start_column + i,
                        cell,
                        format,
                        alignment,
                        font
                    )
        else:
            raise ValueError(f'Invalid value: {direction}')

    def set_matrix(
        self,
        li: list,
        start_row: int = 1,
        start_column: int = 1,
        format = None,
        alignment = None,
        font = None,
        direction: str = 'col',
    ) -> None:
        if direction == 'row':
            for i, row in enumerate(li):
                for j, cell in enumerate(row):
                    self.cell(
                        start_row + i,
                        start_column + j,
                        cell,
                        format,
                        alignment,
                        font
                    )
        elif direction == 'col':
            for i, column in enumerate(li):
                for j, cell in enumerate(column):
                    self.cell(
                        start_row + j,
                        start_column + i,
                        cell,
                        format,
                        alignment,
                        font
                    )
        else:
            raise ValueError(f'Invalid value: {direction}')

    def set_dict(
        self,
        data: dict,
        start_row: int = 1,
        start_column: int = 1,
        *,
        alignment = None,
        font = None
    ) -> None:
        for i, (k, v) in enumerate(data.items()):
            self.cell(
                start_row + i,
                start_column,
                k,
                format="@",
                alignment=alignment,
                font=font,
            )
            if isinstance(v, str):
                self.cell(
                    start_row + i,
                    start_column + 1,
                    v,
                    "@",
                    alignment,
                    font=font,
                )
            elif isinstance(v, Sequence):
                for ii, vv in enumerate(v):
                    self.cell(
                        start_row + i,
                        start_column + ii + 1,
                        str(vv),
                        "@",
                        alignment,
                        font=font,
                    )
            else:
                self.cell(
                    start_row + i,
                    start_column + 1,
                    str(v),
                    "@",
                    alignment=alignment,
                    font=font,
                )