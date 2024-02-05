from __future__ import annotations

import os
import time
import openpyxl as op
from typing import cast, Generator
from workhelper.excel.WorkSheet import WorkSheet
from openpyxl.worksheet.worksheet import Worksheet
import functools

def transfer_WorkSheet(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs) -> WorkSheet:
        ret = func(*args, **kwargs)
        ret.__class__ = WorkSheet
        return ret
    return wrapper

class Excel(op.Workbook):

    def __transfer_WorkSheet(self, ws) -> WorkSheet:
        """转换op.Worksheet为WorkSheet

        Args:
            ws : worksheet

        Returns:
            WorkSheet
        """
        ws.__class__ = WorkSheet
        return cast(WorkSheet, ws)

    @staticmethod
    def load(path: str) -> Excel:
        """读取文件

        Args:
            path (str): 文件路径名

        Returns:
            Excel: _description_

        """
        wb = op.load_workbook(path)
        wb.__class__ = Excel
        return cast(Excel, wb)
    
    @property
    @transfer_WorkSheet
    def active(self) -> WorkSheet:
        """设置当前激活的worksheet"""
        return cast(WorkSheet, super().active)
    
    @transfer_WorkSheet
    def get_sheet_by_name(self, name: str) -> WorkSheet:
        """通过表名获取表格"""
        return cast(WorkSheet, super().get_sheet_by_name(name))
    
    @transfer_WorkSheet
    def __getitem__(self, key: str) -> WorkSheet:
        return cast(WorkSheet, super().__getitem__(key))

    def __iter__(self) -> Generator[WorkSheet, None, None]:
        for ws in super().__iter__():
            yield self.__transfer_WorkSheet(ws)

    @transfer_WorkSheet
    def create_sheet(
        self, title: str | None = None, index: int | None = None
    ) -> WorkSheet:
        """新建表格"""
        return cast(WorkSheet, super().create_sheet(title, index))

    def save_with_stamp(self, path: str, *, open: bool = False) -> None:
        """在表名中增加时间戳后保存"""
        stamp = time.strftime("-%Y_%m_%d-%H_%M_%S", time.localtime())
        path = path[: path.rfind(".")] + stamp + path[path.rfind(".") :]
        self.save(path, open=open)

    def save(self, path: str, *, open: bool = False) -> None:
        """保存"""
        path = os.path.abspath(path)
        super().save(path)
        if open:
            os.startfile(path)

    def has_sheet(self, name: str) -> bool:
        """判断是否存在表格"""
        return name in self.sheetnames